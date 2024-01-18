import sys, argparse
from openpyxl import load_workbook

def main():
    parser = argparse.ArgumentParser('')
    parser.add_argument('--coord')
    args = parser.parse_args();
    x,y = args.coord.split(',')
    print(x,y)
    load_wb = load_workbook("C:/Users/bsa05/workspaces/geom-weather-api/src/main/resources/weather_grid.xlsx")
    sheet = load_wb[0]
    print(sheet.cell(0,0))

main()